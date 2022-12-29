import csv
import os
import pathlib
from typing import List, Dict
import re
import numpy as np
import pandas as pd
import openpyxl
import requests
from matplotlib import pyplot as plt
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from jinja2 import Environment, FileSystemLoader
import pdfkit
from requests.adapters import HTTPAdapter
from urllib3 import Retry
from xlsx2html import xlsx2html
import time
import concurrent.futures



class Vacancy:
    def __init__(self, vacancy: Dict[str, str]):
        self.name = vacancy["name"]
        self.salary = Salary(salary_from=vacancy["salary_from"],
                             salary_to=vacancy["salary_to"],
                             salary_currency=vacancy["salary_currency"],
                             published_at=vacancy["published_at"])
        self.area_name = vacancy["area_name"]
        self.published_at = vacancy["published_at"]
        self.year = self.published_at[:4]

    def get_array_vacancy(self) -> List[str]:
        return [self.name, self.salary.get_average_salary(), self.area_name, self.published_at]

class DataSet:
    def __init__(self, file_name : str):
        self.file_name = file_name
        self.vacancies_objects = self.csv_reader()

    def csv_reader(self) -> (List[Vacancy]):
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file)
            lines = [row for row in file_reader]
            headlines, vacancies = lines[0], lines[1:]
        return self.process_vacancies(headlines, vacancies)

    def process_vacancies(self, headlines: List[str], vacancies: List[List[str]]) -> (List[Vacancy]):
        result = []
        for vacancy in vacancies:
            vacancy = [" ".join(re.sub("<.*?>", "", value).replace('\n', '; ').split()) for value in vacancy]
            result.append(Vacancy({x: y for x, y in zip([r for r in headlines], [v for v in vacancy])}))
        return result

class Salary:
    def __init__(self, salary_from: str or int or float, salary_to: str or int or float, salary_currency: str, published_at: str):
        self.salary_from = self.check_void_value()
        self.salary_to = self.check_void_value()
        self.salary_currency = salary_currency
        self.published_at = published_at
        self.month_year = f"{self.published_at[5:7]}/{self.published_at[:4]}"

    def check_void_value(value: str or int or float) -> float:
        if type(value) == str and value == "":
            return 0
        return float(value)

    def get_average_salary(self):
        return round(((self.salary_from + self.salary_to)
                      * ProcessValutes(self.month_year, self.salary_currency).get_valutes()) / 2, 4)



class SplitCsvFileByYear:
    def __init__(self, file_name: str, directory: str):
        self.file_name = file_name
        self.dir_name = directory
        self.headlines, self.vacancies = self.csv_read()
        self.csv_process(self.headlines, self.vacancies)

    def csv_read(self) -> tuple:
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file)
            lines = [row for row in file_reader]
        return lines[0], lines[1:]

    def csv_process(self, headlines: List[str], vacancies: List[List[str]]) -> None:
        cur_year = "0"
        self.first_vacancy = ""
        os.mkdir(self.dir_name)
        vacancies_cur_year = []
        for vacancy in vacancies:
            if (len(vacancy) == len(headlines)) and ((all([v != "" for v in vacancy])) or (vacancy[1] == "" and vacancy[2] != "") or (vacancy[1] != "" and vacancy[2] == "")):
                vacancy = [" ".join(re.sub("<.*?>", "", value).replace('\n', '; ').split()) for value in vacancy]
                if len(self.first_vacancy) == 0:
                    self.first_vacancy = vacancy
                vacancy_list = [v for v in vacancy]
                if vacancy[-1][:4] != cur_year:
                    if len(vacancies_cur_year) != 0:
                        self.csv_write(headlines, vacancies_cur_year, cur_year)
                        vacancies_cur_year.clear()
                    cur_year = vacancy[-1][:4]
                vacancies_cur_year.append(vacancy_list)
                self.last_vacancy = vacancy
        self.csv_write(headlines, vacancies_cur_year, cur_year)

    def csv_write(self, headlines: List[str], vacancies: List[List[str]], cur_year: str) -> None:
        name = os.path.splitext(self.file_name)
        vacancies = pd.DataFrame(vacancies, columns=headlines)
        vacancies.to_csv(f'{self.dir_name}/{name[0]}_{cur_year}.csv', index=False)

class ProcessValutes:
    def __init__(self, date, salary_currency):
        self.date = date
        self.salary_currency = salary_currency

    def get_valutes(self):
        if self.salary_currency == "RUR":
            return 1
        valutes = pd.read_csv("valutes.csv")
        valute = valutes.loc[valutes["date"] == self.date]
        if valute.__contains__(self.salary_currency):
            return float(valute[self.salary_currency])
        return 0

class YearSalary:
    def __init__(self, param: str, salary: Salary):
        self.param = param
        self.salary = salary.get_average_salary()
        self.count_vacancies = 1

    def add_salary(self, new_salary: Salary):
        self.count_vacancies += 1
        self.salary = self.salary + new_salary.get_average_salary()
        return self

class Report:
    def __init__(self, profession: str, years: List[int], average_salary: List[int],
                 average_salary_profession: List[int], count_vacancies_by_year: List[int],
                 count_vacancies_by_year_prof: List[int], file_name: str):
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.profession = profession
        self.file_name = file_name

    def add_border_and_align(self, worksheet: Worksheet, side: Side, count_columns: int, rows: List[str]) -> None:
        for i in range(1, count_columns):
            for row in rows:
                if i == 1:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                    worksheet[row + str(i)].font = Font(bold=True)
                if worksheet[row + str(i)].internal_value != None:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def make_max_column_width(self, worksheet: Worksheet) -> None:
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value != None:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = len(str(cell.value))
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2

    def generate_excel(self) -> None:
        df = [[self.years[i], self.average_salary[i], self.average_salary_profession[i], self.count_vacancies_by_year[i], self.count_vacancies_by_year_prof[i]] for i in range(len(self.years))]
        df.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"])
        df = pd.DataFrame(df, columns=None)
        with pd.ExcelWriter(self.file_name) as writer:
            df.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
        wb = openpyxl.load_workbook(self.file_name)
        worksheet1 = wb["Статистика по годам"]
        thin = Side(border_style="thin")
        self.add_border_and_align(worksheet1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.make_max_column_width(worksheet1)
        wb.save(self.file_name)

class Graphic:
    def __init__(self, profession: str, years: List[int], average_salary: List[int],
                 average_salary_profession: List[int], count_vacancies_by_year: List[int],
                 count_vacancies_by_year_prof: List[int], file_name : str):
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.profession = profession
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8))
        self.grouped_bar_graph(ax1, "Уровень зарплат по годам", self.average_salary, self.years,
                               self.average_salary_profession, 'средняя з/п', f'з/п {self.profession}')
        self.grouped_bar_graph(ax2, 'Количество вакансий по годам', self.count_vacancies_by_year, self.years,
                               self.count_vacancies_by_year_prof, 'Количество вакансий',
                               f'Количество вакансий {self.profession}')
        plt.tight_layout()
        fig.savefig(file_name)

    def grouped_bar_graph(self, ax, title: str, values_x: List[int], values_y: List[int], values_x2: List[int],
                            label_x: str, label_x2: str) -> None:
        ax.grid(axis='y')
        x = np.arange(len(values_y))
        width = 0.4
        ax.bar(x - width / 2, values_x, width, label=label_x)
        ax.bar(x + width / 2, values_x2, width, label=label_x2)
        ax.set_xticks(x, values_y, rotation=90)
        ax.tick_params(axis="both", labelsize=16)
        ax.set_title(title, fontweight='normal', fontsize=20)
        ax.legend(loc="upper left", fontsize=14)

class InputConnect:
    def __init__(self):
        input_data = []
        for question in ["Введите название csv-файла: ", "Введите название профессии: "]:
            print(question, end="")
            input_data.append(input())
        self.csv_file = input_data[0]
        self.profession = input_data[1]

class PdfConverter:
    def __init__(self, graph_name: str, excel_file: str, profession: str):
        self.graph = graph_name
        self.excel_file = excel_file
        self.prof = profession

    def generate_pdf(self) -> None:
        enviroment = Environment(loader=FileSystemLoader('.'))
        template = enviroment.get_template("pdf_template.html")
        graph_path = os.path.abspath(self.graph)
        out_stream = xlsx2html(self.excel_file, sheet="Статистика по годам")
        out_stream.seek(0)
        pdf_template = template.render({"prof" : self.prof,
                                        "graph": graph_path,
                                        "first_table" : out_stream.read()})
        config = pdfkit.configuration(wkhtmltopdf=r"D:\For PDF python\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, 'report3.pdf', configuration=config, options={"enable-local-file-access": ""})

class GetValutesValues:
    def __init__(self, valutes):
        self.valutes = valutes

    def get_valutes(self, date) -> list:
        session = requests.Session()
        retry = Retry(connect = 3, backoff_factor = 0.5)
        adapter = HTTPAdapter(max_retries=retry)
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        url = f"https://www.cbr.ru/scripts/XML_daily.asp?date_req=01/{date}d=1"
        res = session.get(url)
        cur_df = pd.read_xml(res.text)
        values = []
        for valute in self.valutes:
            if valute in cur_df["CharCode"].values:
                values.append(round(float(cur_df.loc[cur_df["CharCode"] == valute]["Value"].values[0].replace(',', ".")) / float(cur_df.loc[cur_df["CharCode"] == valute]["Nominal"]), 4))
            else:
                values.append(0)
        return [date] + values

    def get_date(first_date, second_date) -> list:
        resultes = []
        for year in range(int(first_date[:4]), int(second_date[:4]) + 1):
            num = 1
            if str(year) == first_date[:4]:
                num = int(first_date[-2:])
            for month in range(num, 13):
                if len(str(month)) == 2:
                    resultes.append(f"{month}/{year}")
                else:
                    resultes.append(f"0{month}/{year}")
                if str(year) == second_date[:4] and (str(month) == second_date[-2:] or f"0{month}" == second_date[-2:]):
                    break
        return resultes

class Statistic:
    def __init__(self, profession: str):
        self.profession = profession

    def process_data(self, file_name: str) -> tuple:
        data = DataSet(file_name).vacancies_objects
        data_profession = [d for d in data if self.profession in d.name]
        year_salary = self.convert_to_param(data)
        professions_year_salary = self.add_missing_years(self.convert_to_param(data_profession), year_salary)
        year_salary, year_vacancy = self.convert_to_dict(year_salary)
        professions_year_salary, professions_year_vacancies = self.convert_to_dict(professions_year_salary)
        return year_salary, year_vacancy, professions_year_salary, professions_year_vacancies

    def convert_to_param(self, vacancies: List[Vacancy]) -> list:
        param_salary = {}
        for vacancy in vacancies:
            if not param_salary.__contains__(vacancy.year):
                param_salary[vacancy.year] = YearSalary(vacancy.year, vacancy.salary)
            else:
                param_salary[vacancy.year] = param_salary[vacancy.year].add_salary(vacancy.salary)
        return [param_salary[d] for d in param_salary]

    def convert_to_dict(self, param_salary: List[YearSalary]) -> tuple:
        return {x: y for x, y in zip([int(r.param) for r in param_salary],
                                     [0 if v.count_vacancies == 0 else int(v.salary / v.count_vacancies) for v in param_salary])},\
               {x: y for x, y in zip([int(r.param) for r in param_salary], [v.count_vacancies for v in param_salary])}

    def add_missing_years(self, param_salary: List[YearSalary], year_salary : List[YearSalary]) -> List[YearSalary]:
        years = [i.param for i in year_salary]
        s_years = [el.param for el in param_salary]
        for y in years:
            if y not in s_years:
                param_salary.insert(int(y) - int(years[0]), YearSalary(y, Salary("0", "0", "RUR", "2003-10-07T00:00:00+0400")))
                param_salary[int(y) - int(years[0])].count_vacancies = 0
        return param_salary

class CreateStatisticFiles:
    def __init__(self, year_salary: Dict[int, int], year_vacancy: Dict[int, int], professions_year_salary: Dict[int, int],
                 professions_year_vacancies: Dict[int, int], profession: str):
        self.year_salary = year_salary
        self.year_vacancy = year_vacancy
        self.professions_year_salary = professions_year_salary
        self.professions_year_vacancies = professions_year_vacancies
        self.profession = profession

    def create_files(self) -> None:
        output_data = {"Динамика уровня зарплат по годам:": self.year_salary,
                       "Динамика количества вакансий по годам:": self.year_vacancy,
                       "Динамика уровня зарплат по годам для выбранной профессии:": self.professions_year_salary,
                       "Динамика количества вакансий по годам для выбранной профессии:": self.professions_year_vacancies}
        [print(i, output_data[i]) for i in output_data]
        excel_file = "report.xlsx"
        report = Report(profession=self.profession,
                        years=[i for i in self.year_salary],
                        average_salary=[self.year_salary[i] for i in self.year_salary],
                        average_salary_profession=[self.professions_year_salary[i] for i in self.professions_year_salary],
                        count_vacancies_by_year=[self.year_vacancy[i] for i in self.year_vacancy],
                        count_vacancies_by_year_prof=[self.professions_year_vacancies[i] for i in self.professions_year_vacancies],
                        file_name=excel_file)
        report.generate_excel()
        graph_name = "graph.png"
        graph = Graphic(profession=self.profession,
                        years=[i for i in self.year_salary],
                        average_salary=[self.year_salary[i] for i in self.year_salary],
                        average_salary_profession=[self.professions_year_salary[i] for i in self.professions_year_salary],
                        count_vacancies_by_year=[self.year_vacancy[i] for i in self.year_vacancy],
                        count_vacancies_by_year_prof=[self.professions_year_vacancies[i] for i in self.professions_year_vacancies],
                        file_name=graph_name)
        pdf = PdfConverter(graph_name=graph_name, excel_file=excel_file, profession=self.profession)
        pdf.generate_pdf()


directory = 'vacancies_by_year'
if __name__ == "__main__":
    year_salary, year_vacancy, professions_year_salary, professions_year_vacancies = {}, {}, {}, {}
    inp = InputConnect()
    spl = SplitCsvFileByYear(inp.csv_file, directory)
    start = time.time()
    files = [str(file) for file in pathlib.Path(f"./{directory}").iterdir()]
    stats = Statistic(inp.profession)
    with concurrent.futures.ProcessPoolExecutor() as executor:
        r = list(executor.map(stats.process_data, files))
        for el in r:
            for i, value in zip(range(4), [year_salary, year_vacancy, professions_year_salary, professions_year_vacancies]):
                value.update(el[i])
    CreateStatisticFiles(year_salary, year_vacancy, professions_year_salary, professions_year_vacancies, inp.profession).create_files()